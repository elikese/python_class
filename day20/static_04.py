from datetime import datetime
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def crawl_ppomppu_page(page_num):
    # 기본 설정
    url = f"https://www.ppomppu.co.kr/zboard/zboard.php?id=ppomppu&page={page_num}&divpage=101"

    print(f"뽐뿌 {page_num}페이지 크롤링 시작...")

    # 웹페이지 요청
    response = requests.get(url)
    print(f"응답 상태코드: {response.status_code}")

    # HTML 파싱
    soup = BeautifulSoup(response.text, "html.parser")

    # 게시글 목록 찾기
    post_rows = soup.find_all("tr", class_="baseList")
    print(f"찾은 게시글 수: {len(post_rows)}개")

    # 데이터 저장할 리스트
    post_list = []

    # 각 게시글 정보 추출
    for i, row in enumerate(post_rows[:21]):
        print(f"--- {i + 1}번째 게시글 처리 중 ---")

        # 썸네일 이미지 추출
        img = row.find("img")
        if img:
            img_url = img.get("src")
            print("이미지: https:" + f"{img_url}")
        else:
            img_url = None
            print("이미지: 없음")

        # 제목, 링크 추출
        title_link = row.find("a", class_="baseList-title")
        if title_link:
            title = title_link.get_text(strip=True)
            link_url = title_link.get("href")
            print(f"제목: {title}")
            print("링크: https://www.ppomppu.co.kr/zboard/" + f"{link_url}")
        else:
            print("제목을 찾을 수 없음")
            continue

        # 작성자 추출
        author_element = row.find("span", class_="baseList-name")
        if author_element:
            author = author_element.get_text(strip=True)
            print(f"작성자: {author}")
        else:
            author = "익명"
            print("작성자: 익명")

        # 시간 추출
        time_element = row.find("time", class_="baseList-time")
        if time_element:
            post_time = time_element.get_text(strip=True)
            print(f"시간: {post_time}")
        else:
            post_time = ""
            print("시간: 정보없음")

        # 데이터 딕셔너리 생성
        post_data = {
            "title": title,
            "author": author,
            "time": post_time,
            "image_url": "https:" + img_url,
            "link_url": "https://www.ppomppu.co.kr/zboard/" + link_url,
        }

        # 리스트에 추가
        post_list.append(post_data)

    # 결과 출력
    print(f"\n" + "=" * 50)
    print(f"{page_num}페이지 크롤링 완료! 총 {len(post_list)}개 게시글 수집")
    print("=" * 50)

    return post_list


# 전체 데이터 저장할 리스트
all_data = []
total_page_num = 3

# 3페이지 크롤링
for page in range(1, total_page_num + 1):
    page_data = crawl_ppomppu_page(page)
    all_data += page_data

    print(f"\n{page}페이지 데이터 추가됨. 현재 총 {len(all_data)}개")

    # 다음 페이지 전에 잠시 대기
    time.sleep(2)

print(f"최종 결과 총 {len(all_data)}개의 게시글 수집 완료!")
print(f"전체 데이터: {all_data}")


def save_to_excel(data, filename):
    """데이터를 엑셀 파일로 저장하는 함수"""
    print(f"엑셀 파일 생성")

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active  # 기본 워크시트 활성화(선택)
    ws.title = "뽐뿌 게시글"  # 시트 이름 변경

    # 헤더 설정
    headers = ["번호", "제목", "작성자", "시간", "이미지URL", "링크URL"]

    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")  # 폰트 색
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 색깔
    header_alignment = Alignment(horizontal="center", vertical="center")  # 가운데 정렬

    # 헤더 추가
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_idx + 1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 데이터 추가
    for row_idx, post in enumerate(data):
        ws.cell(row=row_idx + 2, column=1, value=row_idx + 1)  # 번호
        ws.cell(row=row_idx + 2, column=2, value=post.get("title"))  # 제목
        ws.cell(row=row_idx + 2, column=3, value=post.get("author"))  # 작성자
        ws.cell(row=row_idx + 2, column=4, value=post.get("time"))  # 시간
        ws.cell(row=row_idx + 2, column=5, value=post.get("image_url"))  # 이미지URL
        ws.cell(row=row_idx + 2, column=6, value=post.get("link_url"))  # 링크URL

    # 파일 저장
    wb.save(filename)
    print(f"엑셀 파일 저장 완료: {filename}")


now = datetime.now().strftime("%Y%m%d_%H%M%S")
filename = f"뽐뿌게시글_{now}.xlsx"
save_to_excel(all_data, filename)
