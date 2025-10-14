from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# pip install selenium
# pip install webdriver-manager
# pip install openpyxl

print("네이버 웹툰 크롤링을 시작합니다...")

# 드라이버 설정 및 네이버 접속
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://naver.com")
driver.maximize_window()
sleep(0.5)

# 웹툰 홈으로 이동
webtoon_home_link = driver.find_element(by=By.CSS_SELECTOR, value="#shortcutArea > ul > li:nth-child(9) > a")
webtoon_home_url = webtoon_home_link.get_attribute("href")
driver.get(webtoon_home_url)
sleep(0.5)
print(f"웹툰 홈: {driver.current_url}")

# 웹툰 메뉴 클릭
webtoon_link = driver.find_element(by=By.CSS_SELECTOR, value="#menu > li:nth-child(2) > a")
webtoon_link.click()
sleep(0.5)
print(f"웹툰 페이지: {driver.current_url}")

# 요일별 탭 찾기
day_of_weeks = driver.find_elements(by=By.CSS_SELECTOR, value="#wrap > header > div.SubNavigationBar__snb_wrap--A5gfM > nav > ul > li > a")

# 웹툰 데이터를 저장할 리스트
webtoon_list = []  # 7개의 dict
"""
[
    {
        "day_of_week": "월",
        "webtoon_items": [
            {
                "title": "title1",
                "author": "author1"
            },
            {
                "title": "title2",
                "author": "author2"
            },
            ...
        ]
    },
    {
        "day_of_week": "화",
        "webtoon_items": [
            {
                "title": "title1",
                "author": "author1"
            },
            {
                "title": "title2",
                "author": "author2"
            },
            ...
        ]
    }
]
"""

# 6. 각 요일별로 크롤링 (월~일요일)
for day_of_week in day_of_weeks[1:8]:  # 1~7번째 요소(월~일)
    print(f"\n=== {day_of_week.text}요일 크롤링 시작 ===")

    # 요일 탭 클릭
    day_of_week.click()
    sleep(0.5)

    # 해당 요일 데이터 저장용 딕셔너리
    webtoon_dict = {"day_of_week": day_of_week.text, "webtoon_items": []}

    # 웹툰 아이템들 찾기
    webtoon_items = driver.find_elements(by=By.CSS_SELECTOR, value="#content > div:nth-child(1) > ul > li")

    for idx, webtoon_item in enumerate(webtoon_items):
        try:
            # 요소까지 스크롤
            driver.execute_script("arguments[0].scrollIntoView(true)", webtoon_item)
            sleep(0.2)

            # 이미지 URL 추출
            webtoon_item_img = webtoon_item.find_element(by=By.CSS_SELECTOR, value=".Poster__image--d9XTI")
            webtoon_item_img_src = webtoon_item_img.get_attribute("src")

            # 제목 추출
            webtoon_item_title = webtoon_item.find_element(by=By.CSS_SELECTOR, value=".ContentTitle__title--e3qXt .text")
            webtoon_item_title_text = webtoon_item_title.text

            # 작가 추출
            webtoon_item_author = webtoon_item.find_element(by=By.CSS_SELECTOR, value=".ContentAuthor__author--CTAAP")
            webtoon_item_author_text = webtoon_item_author.text

            # 평점 추출
            webtoon_item_rating = webtoon_item.find_element(by=By.CSS_SELECTOR, value=".Rating__star_area--dFzsb .text")
            webtoon_item_rating_text = webtoon_item_rating.text

            # 웹툰 정보 딕셔너리 생성
            webtoon_item_dict = {
                "img": webtoon_item_img_src,
                "title": webtoon_item_title_text,
                "author": webtoon_item_author_text,
                "rating": webtoon_item_rating_text,
            }

            # 해당 요일 리스트에 추가
            webtoon_dict["webtoon_items"].append(webtoon_item_dict)

            print(f"{idx + 1}. {webtoon_item_title_text} - {webtoon_item_author_text} ({webtoon_item_rating_text})")

            sleep(0.1)

        except Exception as e:
            print(f"  {idx + 1}번째 웹툰 크롤링 실패: {e}")
            continue

    # 요일별 데이터를 전체 리스트에 추가
    webtoon_list.append(webtoon_dict)
    print(f"{day_of_week.text}요일 크롤링 완료: {len(webtoon_dict['webtoon_items'])}개")

# 드라이버 종료
driver.quit()

# 크롤링 완료 메시지
total_webtoon_sum = sum(len(w["webtoon_items"]) for w in webtoon_list)
print(f"\n전체 크롤링 완료! 총 {total_webtoon_sum}개 웹툰 수집")


#######################################################################
# 엑셀 파일 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "웹툰 목록"

# 헤더 설정
headers = ["요일", "제목", "작가", "평점", "이미지URL", "크롤링일시"]
ws.append(headers)

# 헤더 스타일 적용
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers):
    cell = ws.cell(row=1, column=col_num + 1)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 현재 시간
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 데이터를 엑셀에 추가
for webtoon_dict in webtoon_list:
    day_of_week = webtoon_dict["day_of_week"]

    for item in webtoon_dict["webtoon_items"]:
        row_data = [day_of_week, item["title"], item["author"], item["rating"], item["img"], current_time]
        ws.append(row_data)

######################################################### 추가사항 #########################################################
# 요일별 시트 생성
for webtoon_dict in webtoon_list:
    day_of_week = webtoon_dict["day_of_week"]

    # 새 시트 생성
    new_ws = wb.create_sheet(title=f"{day_of_week}요일")

    # 헤더 추가
    sheet_headers = ["순번", "제목", "작가", "평점", "이미지URL"]
    new_ws.append(sheet_headers)  # 그냥 리스트를 한줄로 그대로 넣어 줌

    # 헤더 스타일 적용
    for col_num in range(1, len(sheet_headers) + 1):
        cell = new_ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 추가
    for idx, item in enumerate(webtoon_dict["webtoon_items"]):
        row_data = [idx + 1, item["title"], item["author"], item["rating"], item["img"]]
        new_ws.append(row_data)

    # 열 너비 조정
    new_ws.column_dimensions["A"].width = 8  # 순번
    new_ws.column_dimensions["B"].width = 30  # 제목
    new_ws.column_dimensions["C"].width = 15  # 작가
    new_ws.column_dimensions["D"].width = 10  # 평점
    new_ws.column_dimensions["E"].width = 50  # 이미지URL
######################################################### 추가사항 #########################################################


# 파일명 생성 및 저장
filename = f"네이버웹툰_크롤링결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

# 완료 메시지
print(f"엑셀 파일 저장: {filename}")
