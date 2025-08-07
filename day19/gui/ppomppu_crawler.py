# 크롤링 교육용 - 뽐뿌 게시판 크롤러
# 웹 크롤링의 기본 개념과 실제 구현을 학습할 수 있는 예제

from datetime import datetime  # 현재 날짜/시간을 파일명에 사용
import requests  # HTTP 요청을 보내서 웹페이지 내용 가져오기
from bs4 import BeautifulSoup  # HTML을 파싱(분석)해서 원하는 데이터 추출
import time  # 크롤링 사이의 대기시간 설정 (서버 부하 방지)
from openpyxl import Workbook  # 엑셀 파일 생성
from openpyxl.styles import Font, Alignment  # 엑셀 스타일링


def crawl_ppomppu_page(page_num):
    """
    단일 페이지 크롤링 함수

    Args:
        page_num (int): 크롤링할 페이지 번호 (1, 2, 3...)

    Returns:
        tuple: (데이터 리스트, 에러 메시지)
               성공시: ([{게시글1}, {게시글2}, ...], None)
               실패시: ([], "에러메시지")
    """
    # 1. URL 생성 - f-string을 사용해서 페이지 번호를 동적으로 삽입
    url = f"https://www.ppomppu.co.kr/zboard/zboard.php?id=ppomppu&page={page_num}&divpage=101"

    try:
        # 2. HTTP GET 요청 보내기
        response = requests.get(url, timeout=10)  # 10초 타임아웃 설정
        response.raise_for_status()  # 4xx, 5xx 에러가 있으면 예외 발생

        # 3. HTML 파싱 - BeautifulSoup으로 HTML을 분석 가능한 객체로 변환
        soup = BeautifulSoup(response.text, "html.parser")

        # 4. 게시글 행들 찾기 - CSS 클래스 선택자 사용
        # 뽐뿌 사이트에서 각 게시글은 <tr class="baseList"> 태그로 되어있음
        post_rows = soup.find_all("tr", class_="baseList")

        # 5. 추출한 데이터를 저장할 리스트
        page_data = []

        # 6. 각 게시글 행을 순회하면서 데이터 추출 (최대 21개로 제한)
        for row in post_rows[:21]:

            # 7. 제목과 링크 추출
            # <a class="baseList-title"> 태그에서 제목과 href 속성 가져오기
            title_link = row.find("a", class_="baseList-title")
            if not title_link:  # 제목 링크가 없으면 이 게시글은 건너뛰기
                continue

            # 텍스트 내용 추출 (strip=True로 앞뒤 공백 제거)
            title = title_link.get_text(strip=True)
            # href 속성값 가져오기 (상대 경로)
            link_url = title_link.get("href")

            # 8. 작성자 추출
            # <span class="baseList-name"> 태그에서 작성자명 가져오기
            author_element = row.find("span", class_="baseList-name")
            # 삼항 연산자: 요소가 있으면 텍스트 추출, 없으면 "익명"
            author = author_element.get_text(strip=True) if author_element else "익명"

            # 9. 시간 추출
            # <time class="baseList-time"> 태그에서 게시시간 가져오기
            time_element = row.find("time", class_="baseList-time")
            post_time = time_element.get_text(strip=True) if time_element else ""

            # 10. 이미지 URL 추출
            # <img> 태그의 src 속성에서 이미지 경로 가져오기
            img = row.find("img")
            # 상대경로를 절대경로로 변환 ("//cdn2.ppomppu.co.kr/..." -> "https://cdn2.ppomppu.co.kr/...")
            img_url = "https:" + img.get("src") if img else ""

            # 11. 추출한 데이터를 딕셔너리로 구조화
            post_data = {
                "title": title,
                "author": author,
                "time": post_time,
                "image_url": img_url,
                "link_url": "https://www.ppomppu.co.kr/zboard/" + link_url,  # 절대경로로 변환
            }

            # 12. 페이지 데이터 리스트에 추가
            page_data.append(post_data)

        # 13. 성공시 데이터와 None 반환
        return page_data, None

    except Exception as e:
        # 14. 실패시 빈 리스트와 에러 메시지 반환
        return [], str(e)


def crawl_multiple_pages(page_count, logger=None):
    """
    여러 페이지 크롤링 함수

    Args:
        page_count (int): 크롤링할 총 페이지 수
        logger (function): 진행상황을 출력할 함수 (선택사항)

    Returns:
        tuple: (전체 데이터 리스트, 결과 메시지)
    """
    # 1. 모든 페이지의 데이터를 합칠 리스트
    all_data = []

    # 2. 1페이지부터 지정된 페이지까지 순차적으로 크롤링
    for page in range(1, page_count + 1):

        # 3. 진행상황 로그 출력 (logger 함수가 제공된 경우)
        if logger:
            logger(f"{page}페이지 크롤링 중...")

        # 4. 단일 페이지 크롤링 실행
        page_data, error = crawl_ppomppu_page(page)

        # 5. 에러 발생시 즉시 중단하고 실패 결과 반환
        if error:
            return [], f"{page}페이지 크롤링 실패: {error}"

        # 6. 성공한 페이지 데이터를 전체 데이터에 합치기
        # extend()는 리스트의 각 요소를 개별적으로 추가 (vs append()는 리스트 자체를 하나의 요소로 추가)
        all_data.extend(page_data)

        # 7. 완료 로그 출력
        if logger:
            logger(f"{page}페이지 완료 ({len(page_data)}개 수집)")

        # 8. 서버 부하 방지를 위한 대기시간
        # 마지막 페이지가 아닌 경우에만 대기
        if page < page_count:
            time.sleep(1)  # 1초 대기

    # 9. 모든 페이지 크롤링 완료시 결과 반환
    return all_data, f"총 {len(all_data)}개 게시글 수집 완료"


def save_to_excel(data, filename):
    """
    엑셀 파일로 저장하는 함수

    Args:
        data (list): 저장할 게시글 데이터 리스트
        filename (str): 저장할 파일명

    Returns:
        tuple: (성공여부(bool), 메시지(str))
    """
    # 1. 데이터가 없으면 저장하지 않음
    if not data:
        return False, "저장할 데이터가 없습니다"

    try:
        # 2. 새 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active  # 기본 워크시트 선택
        ws.title = "뽐뿌 게시글"  # 시트명 변경

        # 3. 엑셀 헤더(컬럼명) 정의
        headers = ["번호", "제목", "작성자", "시간", "이미지URL", "링크URL"]

        # 4. 헤더 스타일 설정
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")  # 가운데 정렬

        # 5. 헤더 행 추가 (1행)
        for col_idx, header in enumerate(headers):
            # 엑셀 셀 좌표: (행, 열) - 1부터 시작
            cell = ws.cell(row=1, column=col_idx + 1)
            cell.value = header  # 셀 값 설정
            cell.font = header_font  # 폰트 스타일 적용
            cell.alignment = header_alignment  # 정렬 적용

        # 6. 데이터 행 추가 (2행부터)
        for row_idx, post in enumerate(data):
            # 각 게시글 데이터를 엑셀 행에 추가
            ws.cell(row=row_idx + 2, column=1, value=row_idx + 1)  # 번호 (1부터 시작)
            ws.cell(row=row_idx + 2, column=2, value=post.get("title"))  # 제목
            ws.cell(row=row_idx + 2, column=3, value=post.get("author"))  # 작성자
            ws.cell(row=row_idx + 2, column=4, value=post.get("time"))  # 시간
            ws.cell(row=row_idx + 2, column=5, value=post.get("image_url"))  # 이미지URL
            ws.cell(row=row_idx + 2, column=6, value=post.get("link_url"))  # 링크URL

        # 7. 컬럼 너비 조정 (읽기 편하게)
        ws.column_dimensions["A"].width = 8  # 번호: 좁게
        ws.column_dimensions["B"].width = 50  # 제목: 넓게
        ws.column_dimensions["C"].width = 15  # 작성자: 보통
        ws.column_dimensions["D"].width = 20  # 시간: 보통
        ws.column_dimensions["E"].width = 30  # 이미지URL: 넓게
        ws.column_dimensions["F"].width = 30  # 링크URL: 넓게

        # 9. 파일 저장
        wb.save(filename)
        return (True, f"엑셀 파일 저장 완료: {filename}")

    # 8. 예외 처리 - 구체적인 에러별로 사용자 친화적 메시지 제공
    except PermissionError:
        return (False, "파일이 다른 프로그램에서 열려있어 저장할 수 없습니다. 파일을 닫고 다시 시도해주세요.")
    except FileNotFoundError:
        return (False, "저장할 폴더 경로를 찾을 수 없습니다.")
    except Exception as e:
        return (False, f"예상치 못한 오류: {str(e)}")


def generate_filename():

    # 현재 날짜와 시간을 문자열로 포맷팅
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"뽐뿌게시글_{now}.xlsx"


if __name__ == "__main__":

    # 1페이지 크롤링 테스트
    data, error = crawl_ppomppu_page(1)
    if error:
        print(f"크롤링 실패: {error}")
    else:
        print(f"크롤링 성공: {len(data)}개 게시글")

        # 엑셀 저장 테스트
        filename = generate_filename()
        success, message = save_to_excel(data, filename)
        if success:
            print(f"저장 완료: {filename}")
        else:
            print(f"저장 실패: {message}")
